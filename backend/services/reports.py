"""Report Generation Service — PDF/Excel export for financial reports.

This module provides high-level functions to generate professional PDF and Excel
reports for financial management, inventory, and production analysis.
"""
from datetime import datetime, timedelta, timezone
from io import BytesIO
from typing import Optional, List, Dict, Any, Literal

from sqlalchemy.orm import Session
from sqlalchemy import func

from .. import models
from ..services.financial_events import get_financial_events
from ..services.stock import get_stock_lot_balances

# ReportLab imports for PDF generation
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import mm, cm
from reportlab.lib.enums import TA_LEFT, TA_CENTER, TA_RIGHT
from reportlab.platypus import (
    SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle,
    PageBreak, KeepTogether, Frame, PageTemplate, BaseDocTemplate,
    NextPageTemplate, ListFlowable, ListItem, Image
)
from reportlab.lib.units import inch, cm

# openpyxl for Excel
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, numbers
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows

from .. import models


# ======================================================================
# Styles and Constants
# ======================================================================

# Colors
DARK_BLUE = colors.HexColor("#1a3a5c")
MEDIUM_BLUE = colors.HexColor("#2c5f8a")
LIGHT_BLUE = colors.HexColor("#e8f0fe")
GOLD = colors.HexColor("#d4a843")
LIGHT_GRAY = colors.HexColor("#f5f5f5")
WHITE = colors.white
BLACK = colors.black
RED = colors.HexColor("#dc3545")
GREEN = colors.HexColor("#28a745")
AMBER = colors.HexColor("#ffc107")

# Styles
styles = getSampleStyleSheet()

TITLE_STYLE = ParagraphStyle(
    'ReportTitle',
    parent=styles['Title'],
    fontSize=18,
    leading=22,
    spaceAfter=6,
    textColor=DARK_BLUE,
    alignment=TA_CENTER,
    fontName='Helvetica-Bold'
)

SUBTITLE_STYLE = ParagraphStyle(
    'Subtitle',
    parent=styles['Normal'],
    fontSize=11,
    leading=14,
    spaceAfter=4,
    textColor=MEDIUM_BLUE,
    alignment=TA_CENTER,
    fontName='Helvetica'
)

HEADER_STYLE = ParagraphStyle(
    'ReportHeader',
    parent=styles['Heading2'],
    fontSize=13,
    leading=16,
    spaceBefore=12,
    spaceAfter=6,
    textColor=DARK_BLUE,
    fontName='Helvetica-Bold'
)

NORMAL_STYLE = ParagraphStyle(
    'ReportNormal',
    parent=styles['Normal'],
    fontSize=9,
    leading=12,
    textColor=BLACK,
    fontName='Helvetica'
)

RIGHT_ALIGN = ParagraphStyle(
    'RightAlign',
    parent=styles['Normal'],
    fontSize=9,
    alignment=TA_RIGHT,
    fontName='Helvetica'
)

# Table Styles
TABLE_HEADER_STYLE = TableStyle([
    ('BACKGROUND', (0, 0), (-1, 0), MEDIUM_BLUE),
    ('TEXTCOLOR', (0, 0), (-1, 0), WHITE),
    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
    ('FONTSIZE', (0, 0), (-1, 0), 8),
    ('BOTTOMPADDING', (0, 0), (-1, 0), 6),
    ('TOPPADDING', (0, 0), (-1, 0), 6),
    ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
    ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#cccccc')),
    ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
    ('FONTSIZE', (0, 1), (-1, -1), 7.5),
    ('ROWBACKGROUNDS', (0, 1), (-1, -1), [WHITE, LIGHT_GRAY]),
    ('ALIGN', (1, 1), (-1, -1), 'CENTER'),
    ('ALIGN', (0, 1), (0, -1), 'LEFT'),
)

ALTERNATE_ROW_STYLE = TableStyle([
    ('BACKGROUND', (0, 0), (-1, 0), MEDIUM_BLUE),
    ('TEXTCOLOR', (0, 0), (-1, 0), WHITE),
    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
    ('FONTSIZE', (0, 0), (-1, 0), 8),
    ('BOTTOMPADDING', (0, 0), (-1, 0), 6),
    ('TOPPADDING', (0, 0), (-1, 0), 6),
    ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
    ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#cccccc')),
    ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
    ('FONTSIZE', (0, 1), (-1, -1), 7.5),
    ('ROWBACKGROUNDS', (0, 1), (-1, -1), [WHITE, LIGHT_GRAY]),
    ('ALIGN', (1, 1), (-1, -1), 'CENTER'),
    ('ALIGN', (0, 1), (0, -1), 'LEFT'),
)

# Common Styles for Excel
EXCEL_HEADER_FONT = Font(bold=True, color="FFFFFF", size=10)
EXCEL_HEADER_FILL = PatternFill(start_color="2C5F8A", end_color="2C5F8A", fill_type="solid")
EXCEL_HEADER_ALIGN = Alignment(horizontal='center', vertical='center', wrap_text=True)
EXCEL_DATA_ALIGN = Alignment(horizontal='center', vertical='center', wrap_text=True)
EXCEL_LEFT_ALIGN = Alignment(horizontal='left', vertical='center', wrap_text=True)
EXCEL_RIGHT_ALIGN = Alignment(horizontal='right', vertical='center', wrap_text=True)
EXCEL_THIN_BORDER = Border(
    left=Side(style='thin', color='CCCCCC'),
    right=Side(style='thin', color='CCCCCC'),
    top=Side(style='thin', color='CCCCCC'),
    bottom=Side(style='thin', color='CCCCCC'),
)
EXCEL_HEADER_FILL_ALT = PatternFill(start_color="E8F0FE", end_color="E8F0FE", fill_type="solid")
EXCEL_NUMBER_FORMAT = '#,##0.00'
EXCEL_CURRENCY_FORMAT = '#,##0.00'


# ======================================================================
# Helpers
# ======================================================================

def _format_currency(value: Optional[float]) -> str:
    """Format currency value."""
    if value is None:
        return ""
    return f"{value:,.2f}"


def _format_number(value: Optional[float], decimals: int = 2) -> str:
    """Format number with specified decimals."""
    if value is None:
        return ""
    return f"{value:,.{decimals}f}"


def _safe_float(value: Any) -> float:
    """Safely convert to float."""
    if value is None:
        return 0.0
    try:
        return float(value)
    except (ValueError, TypeError):
        return 0.0


# ======================================================================
# P&L Report (Profit & Loss)
# ======================================================================

def generate_pl_report(
    db: Session,
    owner_id: int,
    start_date: datetime,
    end_date: datetime,
    format: Literal["pdf", "excel"] = "pdf"
) -> BytesIO:
    """
    Generate Profit & Loss report for a date range.
    
    Args:
        db: Database session
        owner_id: Owner ID
        start_date: Report start date
        end_date: Report end date
        format: "pdf" or "excel"
    
    Returns:
        BytesIO buffer with the report
    """
    # Get financial events for the period
    events = db.query(models.FinancialEvent).filter(
        models.FinancialEvent.owner_id == owner_id,
        models.FinancialEvent.event_at >= start_date,
        models.FinancialEvent.event_at <= end_date
    ).all()
    
    # Aggregate by event type
    revenue_ht = 0.0
    revenue_ttc = 0.0
    cogs_ht = 0.0
    cogs_ttc = 0.0
    expense_ht = 0.0
    expense_ttc = 0.0
    waste_cost = 0.0
    
    for event in events:
        if event.event_type == "sale":
            revenue_ht += event.amount_ht or 0
            revenue_ttc += event.amount_ttc or 0
        elif event.event_type == "purchase" and event.event_subtype == "purchase_receive":
            # COGS from purchase receipts
            cogs_ht += event.amount_ht or 0
            cogs_ttc += event.amount_ttc or 0
        elif event.event_type == "production" and event.event_subtype == "production_output":
            # Production output adds to inventory value
            pass
        elif event.event_type == "expense":
            expense_ht += event.amount_ht or 0
            expense_ttc += event.amount_ttc or 0
        elif event.event_type == "waste":
            waste_cost += event.amount_ttc or 0
    
    gross_profit_ht = revenue_ht - cogs_ht
    gross_margin_pct = (gross_profit_ht / revenue_ht * 100) if revenue_ht > 0 else 0
    net_profit_ht = gross_profit_ht - expense_ht - waste_cost
    net_margin_pct = (net_profit_ht / revenue_ht * 100) if revenue_ht > 0 else 0
    
    # Build data structures
    revenue_items = [
        ("Sales Revenue (HT)", revenue_ht),
        ("Sales Revenue (TTC)", revenue_ttc),
    ]
    
    cogs_items = [
        ("COGS (HT)", cogs_ht),
        ("COGS (TTC)", cogs_ttc),
    ]
    
    gross_items = [
        ("Gross Profit (HT)", gross_profit_ht),
        ("Gross Margin %", f"{gross_margin_pct:.1f}%"),
    ]
    
    expense_items = [
        ("Operating Expenses (HT)", expense_ht),
        ("Operating Expenses (TTC)", expense_ttc),
        ("Waste Cost", waste_cost),
    ]
    
    net_items = [
        ("Net Profit (HT)", net_profit_ht),
        ("Net Margin %", f"{net_margin_pct:.1f}%"),
    ]
    
    if format == "pdf":
        return _build_pl_report_pdf(
            start_date, end_date,
            revenue_items, cogs_items, gross_items, expense_items, net_items,
            revenue_ht, cogs_ht, expense_ht, waste_cost
        )
    else:
        return _build_pl_report_excel(
            start_date, end_date,
            revenue_items, cogs_items, gross_items, expense_items, net_items,
            revenue_ht, cogs_ht, expense_ht, waste_cost
        )


def _build_pl_report_pdf(
    start_date: datetime,
    end_date: datetime,
    revenue_items: List[tuple],
    cogs_items: List[tuple],
    gross_items: List[tuple],
    expense_items: List[tuple],
    net_items: List[tuple],
    revenue_ht: float,
    cogs_ht: float,
    expense_ht: float,
    waste_cost: float
) -> BytesIO:
    """Build P&L report as PDF."""
    buffer = BytesIO()
    doc = SimpleDocTemplate(
        BytesIO(),
        pagesize=A4,
        rightMargin=2*cm,
        leftMargin=2*cm,
        topMargin=2*cm,
        bottomMargin=2*cm,
    )
    
    # Build the PDF content
    elements = []
    
    # Title
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Title'],
        fontSize=18,
        leading=22,
        spaceAfter=6,
        textColor=DARK_BLUE,
        alignment=TA_CENTER,
        fontName='Helvetica-Bold'
    )
    
    elements.append(Paragraph("Profit & Loss Statement", TITLE_STYLE))
    elements.append(Paragraph(f"{start_date.strftime('%Y-%m-%d')} to {end_date.strftime('%Y-%m-%d')}", SUBTITLE_STYLE))
    elements.append(Spacer(1, 12))
    
    # Revenue Section
    elements.append(Paragraph("Revenue", HEADER_STYLE))
    revenue_data = [["Description", "Amount (HT)", "Amount (TTC)"]]
    for label, val in revenue_items:
        revenue_data.append([label, _format_currency(val[1]) if isinstance(val, tuple) else _format_currency(val), ""])
    revenue_table = Table(revenue_data, colWidths=[8*cm, 4*cm, 4*cm])
    revenue_table.setStyle(TABLE_HEADER_STYLE)
    elements.append(Table(revenue_data, colWidths=[8*cm, 4*cm, 4*cm]))
    elements.append(Spacer(1, 12))
    
    # COGS
    elements.append(Paragraph("Cost of Goods Sold", HEADER_STYLE))
    cogs_data = [["Description", "Amount (HT)", "Amount (TTC)"]]
    for label, val in cogs_items:
        cogs_data.append([label, _format_currency(val), ""])
    cogs_table = Table(cogs_data, colWidths=[8*cm, 4*cm, 4*cm])
    cogs_table.setStyle(ALTERNATE_ROW_STYLE)
    elements.append(cogs_table)
    elements.append(Spacer(1, 12))
    
    # Gross Profit
    elements.append(Paragraph("Gross Profit", HEADER_STYLE))
    gross_data = [["Description", "Amount (HT)", "Margin %"]]
    for label, val in gross_items:
        if "%" in str(val):
            gross_data.append([label, "", val])
        else:
            gross_data.append([label, _format_currency(val), ""])
    gross_table = Table(gross_data, colWidths=[8*cm, 4*cm, 4*cm])
    gross_table.setStyle(ALTERNATE_ROW_STYLE)
    elements.append(gross_table)
    elements.append(Spacer(1, 12))
    
    # Expenses
    elements.append(Paragraph("Operating Expenses", HEADER_STYLE))
    exp_data = [["Description", "Amount (HT)", "Amount (TTC)"]]
    for label, val in expense_items:
        if isinstance(val, tuple):
            exp_data.append([label, _format_currency(val[0]), _format_currency(val[1])])
        else:
            exp_data.append([label, _format_currency(val), ""])
    exp_table = Table(exp_data, colWidths=[8*cm, 4*cm, 4*cm])
    exp_table.setStyle(ALTERNATE_ROW_STYLE)
    elements.append(exp_table)
    elements.append(Spacer(1, 12))
    
    # Net Profit
    elements.append(Paragraph("Net Profit", HEADER_STYLE))
    net_data = [["Description", "Amount (HT)", "Margin %"]]
    for label, val in net_items:
        if "%" in str(val):
            net_data.append([label, "", val])
        else:
            net_data.append([label, _format_currency(val), ""])
    net_table = Table(net_data, colWidths=[8*cm, 4*cm, 4*cm])
    net_table.setStyle(ALTERNATE_ROW_STYLE)
    elements.append(net_table)
    
    # Build PDF
    doc = SimpleDocTemplate(
        BytesIO(),
        pagesize=A4,
        rightMargin=2*cm,
        leftMargin=2*cm,
        topMargin=2*cm,
        bottomMargin=2*cm,
    )
    doc.build(elements)
    
    # Return BytesIO
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=2*cm, leftMargin=2*cm, topMargin=2*cm, bottomMargin=2*cm)
    doc.build(elements)
    return BytesIO(buffer.getvalue())


def _build_pl_report_excel(
    start_date: datetime,
    end_date: datetime,
    revenue_items: List[tuple],
    cogs_items: List[tuple],
    gross_items: List[tuple],
    expense_items: List[tuple],
    net_items: List[tuple],
    revenue_ht: float,
    cogs_ht: float,
    expense_ht: float,
    waste_cost: float
) -> BytesIO:
    """Build P&L report as Excel workbook."""
    wb = Workbook()
    ws = wb.active
    ws.title = "P&L Report"
    
    # Title
    ws.merge_cells('A1:D1')
    ws['A1'] = "Profit & Loss Statement"
    ws['A1'].font = Font(bold=True, size=16, color="1A3A5C")
    ws['A1'].alignment = Alignment(horizontal='center')
    
    ws.merge_cells('A2:D2')
    ws['A2'] = f"Period: {start_date.strftime('%Y-%m-%d')} to {end_date.strftime('%Y-%m-%d')}"
    ws['A2'].alignment = Alignment(horizontal='center')
    ws['A2'].font = Font(size=11, color="2C5F8A")
    
    # Helper for writing sections
    def write_section(ws, row, title, data, headers=["Description", "Amount (HT)", "Amount (TTC)"]):
        ws.merge_cells(f'A{row}:C{row}')
        ws[f'A{row}'] = row
        ws[f'A{row}'].font = Font(bold=True, size=13, color="1A3A5C")
        row += 1
        
        # Headers
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=1, value=header)
            cell.font = EXCEL_HEADER_FONT
            cell.fill = EXCEL_HEADER_FILL
            cell.alignment = EXCEL_HEADER_ALIGN
            cell.border = EXCEL_THIN_BORDER
        row += 1
        
        for row_data in data:
            if isinstance(row_data[1], tuple):
                ws.cell(row=row, column=1, value=row_data[0])
                ws.cell(row=row, column=2, value=row_data[1][0])
                ws.cell(row=row, column=3, value=row_data[1][1])
            else:
                ws.cell(row=row, column=1, value=row_data[0])
                ws.cell(row=row, column=2, value=row_data[1])
            for col in range(1, 4):
                cell = ws.cell(row=row, column=1)
                cell.border = EXCEL_THIN_BORDER
                cell.alignment = EXCEL_LEFT_ALIGN if col == 1 else EXCEL_RIGHT_ALIGN
                if row_data[1] is not None and isinstance(row_data[1], (int, float)):
                    ws.cell(row=row, column=2).number_format = EXCEL_CURRENCY_FORMAT
                if len(row_data) > 2 and isinstance(row_data[2], (int, float)):
                    ws.cell(row=row, column=3).number_format = EXCEL_CURRENCY_FORMAT
            row += 1
        return row + 1
    
    wb = Workbook()
    ws = wb.active
    ws.title = "P&L Report"
    
    # Title
    ws.merge_cells('A1:D1')
    ws['A1'] = "Profit & Loss Statement"
    ws['A1'].font = Font(bold=True, size=16, color="1A3A5C")
    ws['A1'].alignment = Alignment(horizontal='center')
    
    ws.merge_cells('A2:D2')
    ws['A2'] = f"Period: {start_date.strftime('%Y-%m-%d')} to {end_date.strftime('%Y-%m-%d')}"
    ws['A2'].alignment = Alignment(horizontal='center')
    ws['A2'].font = Font(size=11, color="2C5F8A")
    
    # Revenue
    revenue_data = [(label, val) for label, val in revenue_items]
    row = write_section(ws, 4, "Revenue", revenue_data, ["Description", "HT", "TTC"])
    
    # COGS
    cogs_data = [(label, val) for label, val in cogs_items]
    row = write_section(ws, row, "Cost of Goods Sold", cogs_items, ["Description", "HT", "TTC"])
    
    # Gross Profit
    gross_data = [(label, val) for label, val in gross_items]
    row = write_section(ws, row, "Gross Profit", gross_items, ["Description", "Amount", "Margin %"])
    
    # Expenses
    exp_data = [(label, val) for label, val in expense_items]
    row = write_section(ws, row, "Operating Expenses", expense_items, ["Description", "HT", "TTC"])
    
    # Net Profit
    net_data = [(label, val) for label, val in net_items]
    write_section(ws, row, "Net Profit", net_items, ["Description", "Amount", "Margin %"])
    
    # Save
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return BytesIO(buf.getvalue())


# ======================================================================
# Stock Valuation Report
# ======================================================================

def generate_stock_valuation_report(
    db: Session,
    owner_id: int,
    format: Literal["pdf", "excel"] = "pdf"
) -> BytesIO:
    """Generate stock valuation report by lot/location."""
    from ..services.stock import get_stock_lot_balances
    
    # Get all lot balances
    balances = get_stock_lot_balances(db, owner_id=owner_id, include_zero=False)
    
    # Group by type
    ingredients = []
    semi_finished = []
    products = []
    
    total_ing = 0.0
    total_sf = 0.0
    total_prod = 0.0
    
    for bal in balances:
        lot = bal.lot
        loc = bal.location
        if not lot:
            continue
        uc = lot.unit_cost_snapshot or 0
        val = bal.quantity * uc
        
        row = {
            "type": lot.item_type,
            "name": lot.item_name_snapshot,
            "lot_code": lot.lot_code or "",
            "location": loc.name if loc else "",
            "qty": bal.quantity,
            "unit": lot.unit_snapshot or "",
            "unit_cost": uc,
            "lot_value": val,
            "expires_at": lot.expires_at.isoformat() if lot.expires_at else "",
            "expiry_status": "EXPIRED" if lot.expires_at and lot.expires_at < datetime.now(timezone.utc) else ("EXPIRING" if lot.expires_at and lot.expires_at < datetime.now(timezone.utc) + timedelta(days=7) else "OK")
        }
        
        if lot.item_type == "ingredient":
            ingredients.append(row)
            total_ing += val
        elif lot.item_type == "semi_finished":
            semi_finished.append(row)
            total_sf += val
        elif lot.item_type == "product":
            products.append(row)
            total_prod += val
    
    total = total_ing + total_sf + total_prod
    
    # Expiring soon
    expiring_soon = sum(r["lot_value"] for r in (ingredients + semi_finished + products) 
                        if r["expiry_status"] == "EXPIRING")
    expired = sum(r["lot_value"] for r in (ingredients + semi_finished + products) 
                  if r["expiry_status"] == "EXPIRED")
    
    # Build rows for detail
    rows = []
    for r in [{"type": r["type"], "name": r["name"], "lot_code": r["lot_code"],
               "location": r["location"], "qty": r["qty"], "unit": r["unit"],
               "unit_cost": r["unit_cost"], "lot_value": r["lot_value"],
               "expires_at": r["expires_at"], "expiry_status": r["expiry_status"]}
              for r in ingredients + semi_finished + products]:
        rows.append({
            "type": r["type"],
            "name": r["name"],
            "lot_code": r["lot_code"],
            "location": r["location"],
            "qty": r["qty"],
            "unit": r["unit"],
            "unit_cost": r["unit_cost"],
            "lot_value": r["lot_value"],
            "expires_at": r["expires_at"],
            "expiry_status": r["expiry_status"]
        })
    
    if "pdf" == "pdf":
        return _build_stock_valuation_pdf(
            total_ing=sum_ing, total_sf=total_sf, total_prod=total_prod, total=total_all,
            expiring_soon=expiring_soon, expired=expired, rows=rows
        )
    else:
        return _build_stock_valuation_excel(
            total_ing, total_sf, total_prod, total, expiring_soon, expired, rows
        )


def _build_stock_valuation_pdf(
    total_ing: float, total_sf: float, total_prod: float, total: float,
    expiring_soon: float, expired: float,
    rows: List[dict]
) -> BytesIO:
    """Build stock valuation report as PDF."""
    buffer = BytesIO()
    doc = SimpleDocTemplate(
        BytesIO(),
        pagesize=A4,
        rightMargin=2*cm,
        leftMargin=2*cm,
        topMargin=2*cm,
        bottomMargin=2*cm,
    )
    
    elements = []
    
    # Title
    elements.append(Paragraph("Stock Valuation Report", TITLE_STYLE))
    elements.append(Paragraph(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}", SUBTITLE_STYLE))
    elements.append(Spacer(1, 12))
    
    # Summary
    elements.append(Paragraph("Summary", HEADER_STYLE))
    summary_data = [
        ["Category", "Total Value"],
        ["Ingredients", _format_currency(sum_ing)],
        ["Semi-Finished", _format_currency(total_sf)],
        ["Products", _format_currency(total_prod)],
        ["TOTAL", _format_currency(total)],
        ["Expiring Soon (7 days)", _format_currency(expiring_soon)],
        ["Expired", _format_currency(expired)],
    ]
    summary_table = Table(summary_data, colWidths=[10*cm, 8*cm])
    summary_table.setStyle(ALTERNATE_ROW_STYLE)
    elements.append(summary_table)
    elements.append(Spacer(1, 20))
    
    # Detail table
    elements.append(Paragraph("Detail by Lot", HEADER_STYLE))
    
    headers = ["Type", "Item", "Lot Code", "Location", "Qty", "Unit", "Unit Cost", "Lot Value", "Expiry", "Status"]
    data = [["Type", "Item", "Lot Code", "Location", "Qty", "Unit", "Unit Cost", "Lot Value", "Expiry", "Status"]]
    
    for r in rows:
        data.append([
            r["type"].title(), r["name"], r["lot_code"], r["location"],
            f"{r['qty']:.2f}", r["unit"], f"{r['unit_cost']:.4f}",
            f"{r['lot_value']:.2f}", r["expires_at"][:10] if r["expires_at"] else "",
            r["expiry_status"]
        ])
    
    # Create table with reasonable column widths
    col_widths = [2*cm, 3*cm, 2.5*cm, 2*cm, 1.5*cm, 1.5*cm, 2*cm, 2.5*cm, 2*cm, 1.5*cm]
    table = Table([[Paragraph(str(c), NORMAL_STYLE) for c in row] for row in [["Type", "Item", "Lot Code", "Location", "Qty", "Unit", "Unit Cost", "Lot Value", "Expiry", "Status"]] + 
                   [[r["type"], r["name"], r["lot_code"], r["location"], 
                     f"{r['qty']:.2f}", r["unit"], f"{r['unit_cost']:.4f}",
                      f"{r['lot_value']:.2f}", r["expires_at"][:10] if r["expires_at"] else "", r["expiry_status"]] 
                     for r in rows]], 
                  colWidths=[2*cm, 3*cm, 2.5*cm, 2*cm, 1.5*cm, 1.5*cm, 2*cm, 2.5*cm, 2*cm, 1.5*cm])
    table.setStyle(ALTERNATE_ROW_STYLE)
    
    elements.append(table)
    
    doc = SimpleDocTemplate(BytesIO(), pagesize=A4, rightMargin=2*cm, leftMargin=2*cm, topMargin=2*cm, bottomMargin=2*cm)
    doc.build(elements)
    return BytesIO(buffer.getvalue())


def _build_stock_valuation_excel(
    total_ing: float, total_sf: float, total_prod: float, total: float,
    expiring_soon: float, expired: float,
    rows: List[dict]
) -> BytesIO:
    """Build stock valuation as Excel workbook."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Stock Valuation"
    
    # Title
    ws.merge_cells('A1:F1')
    ws['A1'] = "Stock Valuation Report"
    ws['A1'].font = Font(bold=True, size=16, color="1A3A5C")
    ws['A1'].alignment = Alignment(horizontal='center')
    
    ws.merge_cells('A2:F2')
    ws['A2'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    ws['A2'].alignment = Alignment(horizontal='center')
    ws['A2'].font = Font(size=11, color="2C5F8A")
    
    # Summary
    ws.merge_cells('A4:F4')
    ws['A4'] = "Summary"
    ws['A4'].font = Font(bold=True, size=13, color="1A3A5C")
    
    summary_data = [
        ["Category", "Total Value"],
        ["Ingredients", f"{ing_total:,.2f}"],
        ["Semi-Finished", f"{sf_total:,.2f}"],
        ["Products", f"{prod_total:,.2f}"],
        ["TOTAL", f"{total:,.2f}"],
        ["Expiring Soon (7 days)", f"{expiring_soon:,.2f}"],
        ["Expired", f"{expired:,.2f}"],
    ]
    
    for i, row_data in enumerate(summary_data, start=5):
        ws[f'A{i}'] = row_data[0]
        ws[f'B{i}'] = row_data[1]
        for cell in [f'A{i}', f'B{i}']:
            ws[cell].border = EXCEL_THIN_BORDER
            ws[cell].alignment = EXCEL_LEFT_ALIGN if cell.startswith('A') else EXCEL_RIGHT_ALIGN
            if i > 5:
                ws[cell].font = Font(bold=True)
            if cell.startswith('B') and i > 5:
                ws[cell].number_format = '#,##0.00'
    
    # Detail
    row = 14
    ws[f'A{14}'] = "Detail by Lot"
    ws[f'A{14}'].font = Font(bold=True, size=13, color="1A3A5C")
    
    headers = ["Type", "Item", "Lot Code", "Location", "Qty", "Unit", "Unit Cost", "Lot Value", "Expiry", "Status"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=13, column=col, value=header)
        cell.font = EXCEL_HEADER_FONT
        cell.fill = EXCEL_HEADER_FILL
        cell.alignment = EXCEL_HEADER_ALIGN
        cell.border = EXCEL_THIN_BORDER
    
    for i, r in enumerate(rows, start=15):
        data = [r["type"], r["name"], r["lot_code"], r["location"],
                r["qty"], r["unit"], r["unit_cost"], r["lot_value"],
                r["expires_at"][:10] if r["expires_at"] else "", r["expiry_status"]]
        for col, val in enumerate(data, 1):
            cell = ws.cell(row=13+i, column=col, value=val)
            cell.border = EXCEL_THIN_BORDER
            cell.alignment = EXCEL_LEFT_ALIGN if col <= 4 else (EXCEL_RIGHT_ALIGN if col in [5,7] else EXCEL_LEFT_ALIGN)
            if col == 7:
                cell.number_format = '#,##0.00'
            elif col == 8:
                cell.number_format = '#,##0.00'
    
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return BytesIO(buf.getvalue())


# ======================================================================
# Production Cost Analysis Report
# ======================================================================

def generate_production_cost_report(
    db: Session,
    owner_id: int,
    start_date: datetime,
    end_date: datetime,
    format: Literal["pdf", "excel"] = "pdf"
) -> BytesIO:
    """Generate production cost analysis report."""
    # Get production batches in range
    batches = db.query(models.ProductionBatch).filter(
        models.ProductionBatch.owner_id == owner_id,
        models.ProductionBatch.completed_at >= start_date,
        models.ProductionBatch.completed_at <= end_date,
        models.ProductionBatch.stage == "ready"
    ).all()
    
    rows = []
    total_revenue = 0.0
    total_cost = 0.0
    
    for batch in batches:
        if not batch.cost_snapshot:
            continue
        
        cost_snapshot = batch.cost_snapshot
        revenue = batch.quantity * batch.product.price if batch.product else 0
        cost = cost_snapshot.get("total_cost", 0) if cost_snapshot else 0
        margin = revenue - cost if revenue > 0 else 0
        margin_pct = (margin / revenue * 100) if revenue > 0 else 0
        
        rows.append({
            "batch_id": batch.id,
            "product": batch.product.name if batch.product else batch.product_id,
            "quantity": batch.quantity,
            "planned_date": batch.planned_for_date,
            "completed": batch.completed_at.strftime("%Y-%m-%d") if batch.completed_at else "",
            "stage": batch.stage,
            "cost_per_unit": cost_snapshot.get("cost_per_unit", 0) if cost_snapshot else 0,
            "total_cost": cost,
            "revenue": revenue,
            "margin": margin,
            "margin_pct": margin_pct,
        })
        
        total_revenue += revenue
        total_cost += cost
    
    total_margin = total_revenue - total_cost
    total_margin_pct = (total_margin / total_revenue * 100) if total_revenue > 0 else 0
    
    if "pdf" == "pdf":
        return _build_production_cost_pdf(start_date, end_date, rows, total_revenue, total_cost, total_margin, total_margin_pct)
    else:
        return _build_production_cost_excel(start_date, end_date, rows, total_revenue, total_cost, total_margin, total_margin_pct)


def _build_production_cost_pdf(start_date, end_date, rows, total_revenue, total_cost, total_margin, total_margin_pct):
    """Build production cost report as PDF."""
    buffer = BytesIO()
    doc = SimpleDocTemplate(
        BytesIO(),
        pagesize=A4,
        rightMargin=2*cm,
        leftMargin=2*cm,
        topMargin=2*cm,
        bottomMargin=2*cm,
    )
    
    elements = []
    
    elements.append(Paragraph("Production Cost Analysis", TITLE_STYLE))
    elements.append(Paragraph(f"Period: {start_date.strftime('%Y-%m-%d')} to {end_date.strftime('%Y-%m-%d')}", SUBTITLE_STYLE))
    elements.append(Spacer(1, 12))
    
    # Summary
    summary_data = [
        ["Metric", "Value"],
        ["Total Batches", str(len(batches))],
        ["Total Revenue", f"{total_revenue:,.2f}"],
        ["Total Cost", f"{total_cost:,.2f}"],
        ["Total Margin", f"{total_margin:,.2f}"],
        ["Margin %", f"{total_margin_pct:.1f}%"],
    ]
    summary_table = Table(summary_data, colWidths=[8*cm, 8*cm])
    summary_table.setStyle(ALTERNATE_ROW_STYLE)
    elements.append(Paragraph("Summary", HEADER_STYLE))
    elements.append(summary_table)
    elements.append(Spacer(1, 20))
    
    # Detail table
    headers = ["Batch ID", "Product", "Qty", "Planned Date", "Completed", "Stage", "Cost/Unit", "Total Cost", "Revenue", "Margin", "Margin %"]
    data = [["Batch ID", "Product", "Qty", "Planned", "Completed", "Stage", "Cost/Unit", "Total Cost", "Revenue", "Margin", "Margin %"]]
    
    for r in rows:
        data.append([
            r["batch_id"], r["product"], str(r["quantity"]), r["planned_date"],
            r["completed"], r["stage"], f"{r['cost_per_unit']:.4f}",
            f"{r['total_cost']:,.2f}", f"{r['revenue']:,.2f}",
            f"{r['margin']:,.2f}", f"{r['margin_pct']:.1f}%"
        ])
    
    table = Table(
        [["Batch ID", "Product", "Qty", "Planned", "Completed", "Stage", "Cost/Unit", "Total Cost", "Revenue", "Margin", "Margin %"]] + 
        [[r["batch_id"], r["product"], str(r["quantity"]), r["planned_date"], r["completed"], r["stage"],
          f"{r['cost_per_unit']:.4f}", f"{r['total_cost']:,.2f}", f"{r['revenue']:,.2f}", f"{r['margin']:,.2f}", f"{r['margin_pct']:.1f}%"] for r in rows],
        colWidths=[2*cm, 3*cm, 1.5*cm, 2*cm, 2*cm, 1.5*cm, 2*cm, 2.5*cm, 2.5*cm, 2.5*cm, 2*cm]
    )
    table.setStyle(ALTERNATE_ROW_STYLE)
    
    elements.append(Paragraph("Batch Details", HEADER_STYLE))
    elements.append(table)
    
    doc = SimpleDocTemplate(BytesIO(), pagesize=A4, rightMargin=2*cm, leftMargin=2*cm, topMargin=2*cm, bottomMargin=2*cm)
    doc.build(elements)
    return BytesIO(buffer.getvalue())


def _build_production_cost_excel(
    start_date, end_date, rows, total_revenue, total_cost, total_margin, total_margin_pct
) -> BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "Production Cost Analysis"
    
    # Title
    ws.merge_cells('A1:K1')
    ws['A1'] = "Production Cost Analysis"
    ws['A1'].font = Font(bold=True, size=16, color="1A3A5C")
    ws['A1'].alignment = Alignment(horizontal='center')
    
    ws.merge_cells('A2:K2')
    ws['A2'] = f"Period: {start_date.strftime('%Y-%m-%d')} to {end_date.strftime('%Y-%m-%d')}"
    ws['A2'].alignment = Alignment(horizontal='center')
    ws['A2'].font = Font(size=11, color="2C5F8A")
    
    # Summary
    row = 4
    ws.merge_cells(f'A{row}:B{row}')
    ws[f'A{row}'] = "Summary"
    ws[f'A{row}'].font = Font(bold=True, size=13, color="1A3A5C")
    
    summary = [
        ["Total Batches", len(rows)],
        ["Total Revenue", total_revenue],
        ["Total Cost", total_cost],
        ["Total Margin", total_margin],
        ["Margin %", f"{total_margin_pct:.1f}%"],
    ]
    for i, (label, val) in enumerate(summary, start=5):
        ws[f'A{i}'] = label
        ws[f'B{i}'] = val
        ws[f'A{i}'].font = Font(bold=True)
        ws[f'B{i}'].number_format = '#,##0.00' if isinstance(val, (int, float)) else '@'
    
    # Detail table
    row = 12
    headers = ["Batch ID", "Product", "Qty", "Planned Date", "Completed", "Stage", "Cost/Unit", "Total Cost", "Revenue", "Margin", "Margin %"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=13, column=col, value=header)
        cell.font = EXCEL_HEADER_FONT
        cell.fill = EXCEL_HEADER_FILL
        cell.alignment = EXCEL_HEADER_ALIGN
        cell.border = EXCEL_THIN_BORDER
    
    for i, r in enumerate(rows, start=14):
        ws.cell(row=i, column=1, value=r["batch_id"]).border = EXCEL_THIN_BORDER
        ws.cell(row=i, column=2, value=r["product"]).border = EXCEL_THIN_BORDER
        ws.cell(row=i, column=3, value=r["quantity"]).border = EXCEL_THIN_BORDER
        ws.cell(row=i, column=4, value=r["planned_date"]).border = EXCEL_THIN_BORDER
        ws.cell(row=i, column=5, value=r["completed"]).border = EXCEL_THIN_BORDER
        ws.cell(row=i, column=6, value=r["stage"]).border = EXCEL_THIN_BORDER
        ws.cell(row=i, column=6, value=r["stage"]).border = EXCEL_THIN_BORDER
        ws.cell(row=i, column=7, value=r["cost_per_unit"]).border = EXCEL_THIN_BORDER
        ws.cell(row=i, column=7, value=r["cost_per_unit"]).number_format = '#,##0.0000'
        ws.cell(row=i, column=8, value=r["total_cost"]).border = EXCEL_THIN_BORDER
        ws.cell(row=i, column=8, value=r["total_cost"]).number_format = '#,##0.00'
        ws.cell(row=i, column=9, value=r["revenue"]).border = EXCEL_THIN_BORDER
        ws.cell(row=i, column=9, value=r["revenue"]).number_format = '#,##0.00'
        ws.cell(row=i, column=10, value=r["margin"]).border = EXCEL_THIN_BORDER
        ws.cell(row=i, column=10, value=r["margin"]).number_format = '#,##0.00'
        ws.cell(row=i, column=11, value=r["margin_pct"]).border = EXCEL_THIN_BORDER
        ws.cell(row=i, column=11, value=r["margin_pct"]).number_format = '0.00%'
    
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return BytesIO(buf.getvalue())


# ======================================================================
# Stock Valuation Report
# ======================================================================

def generate_stock_valuation_report(
    db: Session,
    owner_id: int,
    format: Literal["pdf", "excel"] = "pdf"
) -> BytesIO:
    """Generate stock valuation report by lot/location."""
    from ..services.stock import get_stock_lot_balances
    
    # Get all lot balances
    balances = get_stock_lot_balances(db, owner_id=owner_id)
    
    # Group by type
    ingredients = []
    semi_finished = []
    products = []
    
    total_ing = 0.0
    total_sf = 0.0
    total_prod = 0.0
    
    rows = []
    
    for bal in balances:
        lot = bal.lot
        loc = bal.location
        if not lot:
            continue
        uc = lot.unit_cost_snapshot or 0
        val = bal.quantity * uc
        
        row = {
            "type": lot.item_type,
            "name": lot.item_name_snapshot,
            "lot_code": lot.lot_code or "",
            "location": loc.name if loc else "",
            "qty": bal.quantity,
            "unit": lot.unit_snapshot or "",
            "unit_cost": uc,
            "lot_value": val,
            "expires_at": lot.expires_at.isoformat() if lot.expires_at else "",
            "expiry_status": "EXPIRED" if lot.expires_at and lot.expires_at < datetime.now(timezone.utc) else ("EXPIRING" if lot.expires_at and lot.expires_at < datetime.now(timezone.utc) + timedelta(days=7) else "OK")
        }
        
        if lot.item_type == "ingredient":
            ingredients.append(row)
            total_ing += val
        elif lot.item_type == "semi_finished":
            semi_finished.append(row)
            total_sf += val
        elif lot.item_type == "product":
            products.append(row)
            total_prod += val
        
        rows.append({
            "type": lot.item_type,
            "name": lot.item_name_snapshot,
            "lot_code": lot.lot_code or "",
            "location": loc.name if loc else "",
            "qty": bal.quantity,
            "unit": lot.unit_snapshot or "",
            "unit_cost": uc,
            "lot_value": val,
            "expires_at": lot.expires_at.isoformat() if lot.expires_at else "",
            "expiry_status": "EXPIRED" if lot.expires_at and lot.expires_at < datetime.now(timezone.utc) else ("EXPIRING" if lot.expires_at and lot.expires_at < datetime.now(timezone.utc) + timedelta(days=7) else "OK")
        })
    
    total_ing = sum(r["lot_value"] for r in rows if r["type"] == "ingredient")
    total_sf = sum(r["lot_value"] for r in rows if r["type"] == "semi_finished")
    total_prod = sum(r["lot_value"] for r in rows if r["type"] == "product")
    total = total_ing + total_sf + total_prod
    
    expiring_soon = sum(r["lot_value"] for r in rows if r["expiry_status"] == "EXPIRING")
    expired = sum(r["lot_value"] for r in rows if r["expiry_status"] == "EXPIRED")
    
    if "pdf" == "pdf":
        return _build_stock_valuation_pdf(
            total_ing=total_ing, total_sf=total_sf, total_prod=total_prod, total=total_ing+total_sf+total_fp,
            expiring_soon=expiring_soon, expired=expired, rows=rows
        )
    else:
        return _build_stock_valuation_excel(
            total_ing, total_sf, total_prod, sum(total_ing, total_sf, total_fp),
            expiring_soon, expired, rows
        )


def _build_stock_valuation_pdf(
    total_ing: float, total_sf: float, total_prod: float, total: float,
    expiring_soon: float, expired: float,
    rows: List[dict]
) -> BytesIO:
    """Build stock valuation report as PDF."""
    buffer = BytesIO()
    doc = SimpleDocTemplate(
        BytesIO(),
        pagesize=A4,
        rightMargin=2*cm,
        leftMargin=2*cm,
        topMargin=2*cm,
        bottomMargin=2*cm,
    )
    
    elements = []
    
    elements.append(Paragraph("Stock Valuation Report", TITLE_STYLE))
    elements.append(Paragraph(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}", SUBTITLE_STYLE))
    elements.append(Spacer(1, 12))
    
    # Summary
    elements.append(Paragraph("Summary", HEADER_STYLE))
    summary_data = [
        ["Category", "Total Value"],
        ["Ingredients", _format_currency(sum_ing)],
        ["Semi-Finished", _format_currency(sf_total)],
        ["Products", _format_currency(prod_total)],
        ["TOTAL", _format_currency(total)],
        ["Expiring Soon (<7 days)", _format_currency(expiring_soon)],
        ["Expired", _format_currency(expired)],
    ]
    summary_table = Table(summary_data, colWidths=[10*cm, 8*cm])
    summary_table.setStyle(ALTERNATE_ROW_STYLE)
    elements.append(summary_table)
    elements.append(Spacer(1, 20))
    
    # Detail table
    elements.append(Paragraph("Detail by Lot", HEADER_STYLE))
    
    headers = ["Type", "Item", "Lot Code", "Location", "Qty", "Unit", "Unit Cost", "Lot Value", "Expiry", "Status"]
    data = [["Type", "Item", "Lot Code", "Location", "Qty", "Unit", "Unit Cost", "Lot Value", "Expiry", "Status"]]
    
    for r in rows:
        data.append([
            r["type"], r["name"], r["lot_code"], r["location"],
            f"{r['qty']:.2f}", r["unit"], f"{r['unit_cost']:.4f}",
            f"{r['lot_value']:.2f}", r["expires_at"][:10] if r["expires_at"] else "",
            r["expiry_status"]
        ])
    
    table = Table(
        [[Paragraph(str(c), NORMAL_STYLE) for c in row] for row in [["Type", "Item", "Lot Code", "Location", "Qty", "Unit", "Unit Cost", "Lot Value", "Expiry", "Status"]] + 
         [[r["type"], r["name"], r["lot_code"], r["location"], f"{r['qty']:.2f}", r["unit"], f"{r['unit_cost']:.4f}", f"{r['lot_value']:.2f}", r["expires_at"][:10] if r["expires_at"] else "", r["expiry_status"]] for r in rows]],
        colWidths=[2*cm, 3*cm, 2.5*cm, 2*cm, 1.5*cm, 1.5*cm, 2*cm, 2.5*cm, 2*cm, 1.5*cm]
    )
    table.setStyle(ALTERNATE_ROW_STYLE)
    
    elements.append(table)
    
    doc = SimpleDocTemplate(BytesIO(), pagesize=A4, rightMargin=2*cm, leftMargin=2*cm, topMargin=2*cm, bottomMargin=2*cm)
    doc.build(elements)
    return BytesIO(buffer.getvalue())


def _build_stock_valuation_excel(
    total_ing: float, total_sf: float, total_prod: float, total: float,
    expiring_soon: float, expired: float,
    rows: List[dict]
) -> BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "Stock Valuation"
    
    # Title
    ws.merge_cells('A1:F1')
    ws['A1'] = "Stock Valuation Report"
    ws['A1'].font = Font(bold=True, size=16, color="1A3A5C")
    ws['A1'].alignment = Alignment(horizontal='center')
    
    ws.merge_cells('A2:F2')
    ws['A2'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    ws['A2'].alignment = Alignment(horizontal='center')
    ws['A2'].font = Font(size=11, color="2C5F8A")
    
    # Summary
    ws.merge_cells('A4:F4')
    ws['A4'] = "Summary"
    ws['A4'].font = Font(bold=True, size=13, color="1A3A5C")
    
    summary_data = [
        ["Category", "Total Value"],
        ["Ingredients", f"{ing_total:,.2f}"],
        ["Semi-Finished", f"{sf_total:,.2f}"],
        ["Products", f"{prod_total:,.2f}"],
        ["TOTAL", f"{total:,.2f}"],
        ["Expiring Soon (7 days)", f"{expiring_soon:,.2f}"],
        ["Expired", f"{expired:,.2f}"],
    ]
    
    for i, row_data in enumerate(summary_data, start=5):
        ws.cell(row=i, column=1, value=row_data[0])
        ws.cell(row=i, column=2, value=row_data[1])
        for cell in [f'A{i}', f'B{i}']:
            ws[cell].border = EXCEL_THIN_BORDER
            ws[cell].alignment = EXCEL_LEFT_ALIGN if cell.startswith('A') else EXCEL_RIGHT_ALIGN
            if i > 5:
                ws[cell].font = Font(bold=True)
            if cell.startswith('B') and i > 5:
                ws[cell].number_format = '#,##0.00'
    
    # Detail
    row = 14
    ws[f'A{14}'] = "Detail by Lot"
    ws[f'A{14}'].font = Font(bold=True, size=13, color="1A3A5C")
    
    headers = ["Type", "Item", "Lot Code", "Location", "Qty", "Unit", "Unit Cost", "Lot Value", "Expiry", "Status"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=13, column=col, value=header)
        cell.font = EXCEL_HEADER_FONT
        cell.fill = EXCEL_HEADER_FILL
        cell.alignment = EXCEL_HEADER_ALIGN
        cell.border = EXCEL_THIN_BORDER
    
    for i, r in enumerate(rows, start=15):
        data = [r["type"], r["name"], r["lot_code"], r["location"],
                r["qty"], r["unit"], r["unit_cost"], r["lot_value"],
                r["expires_at"][:10] if r["expires_at"] else "", r["expiry_status"]]
        for col, val in enumerate(data, 1):
            cell = ws.cell(row=13+i, column=col, value=val)
            cell.border = EXCEL_THIN_BORDER
            cell.alignment = EXCEL_LEFT_ALIGN if col <= 4 else (EXCEL_RIGHT_ALIGN if col in [5,7] else EXCEL_LEFT_ALIGN)
            if col == 7:
                cell.number_format = '#,##0.00'
            elif col == 8:
                cell.number_format = '#,##0.00'
    
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return BytesIO(buf.getvalue())


# ======================================================================
# Main Report Generator Function
# ======================================================================

def generate_report(
    db: Session,
    owner_id: int,
    report_type: Literal["pl", "stock_valuation", "production_cost", "purchase_aging"],
    format: Literal["pdf", "excel"],
    start_date: Optional[datetime] = None,
    end_date: Optional[datetime] = None,
    as_of: Optional[datetime] = None,
) -> BytesIO:
    """
    Main entry point for report generation.
    
    Args:
        db: Database session
        owner_id: Owner ID
        report_type: Type of report ("pl", "stock_valuation", "production_cost", "purchase_aging")
        format: "pdf" or "excel"
        start_date: Start date for period reports
        end_date: End date for period reports
        as_of: As-of date for point-in-time reports
    
    Returns:
        BytesIO buffer with the report
    """
    if report_type == "pl":
        if not start_date or not end_date:
            raise ValueError("start_date and end_date required for P&L report")
        return generate_pl_report(db, owner_id, start_date, end_date, format)
    
    elif report_type == "stock_valuation":
        return generate_stock_valuation_report(db, owner_id, format)
    
    elif report_type == "production_cost":
        if not start_date or not end_date:
            raise ValueError("start_date and end_date required for production cost report")
        return generate_production_cost_report(db, owner_id, start_date, end_date, format)
    
    elif report_type == "purchase_aging":
        return generate_purchase_aging_report(db, owner_id, end_date or datetime.now(timezone.utc), format)
    
    else:
        raise ValueError(f"Unknown report type: {report_type}")